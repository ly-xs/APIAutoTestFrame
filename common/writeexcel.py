import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(__file__)))
import shutil
import configparser
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from config.config import SOURCE_FILE, TARGET_FILE, CONFIG_FILE

# --------- 读取config.ini配置文件 ---------------
config = configparser.ConfigParser()
config.read(CONFIG_FILE, encoding='UTF-8')
name = config.get("tester", "name")


class ExcelWriter:
    """文件写入数据"""
    GREEN = '00FF00'  # 绿色
    RED = 'FF0000'  # 红色
    DARKYELLOW = '808000'  # 深黄色

    def __init__(self, filename):
        """
        初始化 WriteExcel 类
        :param filename: 文件名
        """
        self.filename = filename
        self._prepare_file()

    def _prepare_file(self):
        """
        准备文件：如果文件不存在，则拷贝模板文件至指定报告目录下
        """
        if not os.path.exists(self.filename):
            shutil.copyfile(SOURCE_FILE, TARGET_FILE)
        self.workbook = load_workbook(self.filename)
        self.sheet = self.workbook.active

    def write_data(self, row, value):
        """
        写入测试结果
        :param row: 数据所在行数
        :param value: 测试结果值
        """
        font_green = Font(name='宋体', color=self.GREEN, bold=True)
        font_red = Font(name='宋体', color=self.RED, bold=True)
        font_darkyellow = Font(name='宋体', color=self.DARKYELLOW, bold=True)
        align = Alignment(horizontal='center', vertical='center')

        column_l = get_column_letter(12)  # 列L
        column_m = get_column_letter(13)  # 列M
        cell_l = f"{column_l}{row}"
        cell_m = f"{column_m}{row}"

        if value == "PASS":
            self.sheet[cell_l] = value
            self.sheet[cell_l].font = font_green
        elif value == "FAIL":
            self.sheet[cell_l] = value
            self.sheet[cell_l].font = font_red

        self.sheet[cell_m] = name
        self.sheet[cell_l].alignment = align
        self.sheet[cell_m].font = font_darkyellow
        self.sheet[cell_m].alignment = align

        self.workbook.save(self.filename)
