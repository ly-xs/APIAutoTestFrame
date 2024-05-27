from openpyxl import load_workbook


class ExcelReader:
    def __init__(self, file_path):
        """
        初始化ExcelReader对象，接受Excel文件的路径。

        :param file_path: str, Excel文件的路径
        """
        self.file_path = file_path
        self.workbook = None
        self.sheet = None

    def load_data(self, sheet_name=None):
        """
        加载Excel文件中的数据。

        :param sheet_name: str, 要读取的工作表名称，默认读取活动工作表
        """
        try:
            self.workbook = load_workbook(self.file_path)
            if sheet_name:
                self.sheet = self.workbook[sheet_name]
            else:
                self.sheet = self.workbook.active
            print(f"数据成功加载自工作表: {self.sheet.title}")
        except Exception as e:
            print(f"加载数据失败: {e}")

    def get_cell_value(self, row, column):
        """
        获取指定单元格的数据。

        :param row: int, 行索引（从1开始）
        :param column: int, 列索引（从1开始）
        :return: 指定单元格的数据
        """
        if self.sheet:
            try:
                return self.sheet.cell(row=row, column=column).value
            except Exception as e:
                print(f"获取单元格数据失败: {e}")
                return None
        else:
            print("没有加载工作表，请先加载数据。")
            return None

    def get_all_data(self):
        """
        获取整个工作表的数据。

        :return: 包含整个工作表数据的二维列表
        """
        if self.sheet:
            try:
                # return [list(row) for row in self.sheet.iter_rows(values_only=True)
                #         if any(cell is not None for cell in row)]
                raw_data = [list(row) for row in self.sheet.iter_rows(values_only=True)
                            if any(cell is not None for cell in row)]
                header = raw_data[0]  # 第一行作为键
                values = raw_data[1:]  # 其他行作为值

                # 将值为None的元素替换为字符串为空
                values = [[cell if cell else "" for cell in row] for row in values]

                return [{header[i]: row[i] for i in range(len(header))} for row in values]
            except Exception as e:
                print(f"获取全部数据失败: {e}")
                return None
        else:
            print("没有加载工作表，请先加载数据。")
            return None


if __name__ == '__main__':
    data = ExcelReader('../database/DemoAPITestCase.xlsx')
    data.load_data()
    print(data.get_all_data())
